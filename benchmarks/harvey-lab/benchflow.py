"""Harvey LAB → BenchFlow converter.

Translates Harvey LAB tasks into BenchFlow task format. Each Harvey LAB
task contains instructions, documents, and rubric criteria graded by an
LLM judge. This converter maps those to either legacy task.toml /
instruction.md / environment / tests structure, or native task.md packages.

Usage:
    # Generate all tasks (native task.md packages by default)
    python benchmarks/harvey-lab/benchflow.py --output-dir /tmp/harvey-lab-tasks

    # Generate legacy split packages (task.toml / instruction.md / tests/)
    python benchmarks/harvey-lab/benchflow.py --output-dir /tmp/harvey-lab-legacy \
        --task-format legacy

    # Generate a subset
    python benchmarks/harvey-lab/benchflow.py --output-dir /tmp/harvey-lab-tasks --limit 10

    # Generate specific tasks
    python benchmarks/harvey-lab/benchflow.py --output-dir /tmp/harvey-lab-tasks \
        --task-ids "corporate-ma/analyze-cim-deal-teaser/scenario-01"
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
import textwrap
from pathlib import Path
from typing import Any

import yaml

_REPO_SRC = Path(__file__).resolve().parents[2] / "src"
_repo_src_path = str(_REPO_SRC)
if _repo_src_path in sys.path:
    sys.path.remove(_repo_src_path)
sys.path.insert(0, _repo_src_path)

from benchflow.task.document import render_task_md  # noqa: E402
from benchflow.task.output_format import (  # noqa: E402
    TASK_OUTPUT_FORMATS,
    TaskOutputFormat,
    ensure_existing_task_output_format,
    validate_task_output_format,
    verifier_dir_name,
)

# Harvey LAB repo root — resolved relative to this script's location.
_SCRIPT_DIR = Path(__file__).resolve().parent
_DEFAULT_HARVEY_ROOT = _SCRIPT_DIR.parent.parent.parent / "harvey-labs"

# Difficulty heuristic thresholds (by criteria count).
_EASY_THRESHOLD = 35
_HARD_THRESHOLD = 80
TASK_FORMATS = TASK_OUTPUT_FORMATS
TaskFormat = TaskOutputFormat


def _sanitize_name(raw: str) -> str:
    """Lowercase, replace non-alphanumeric with hyphens, collapse runs."""
    import re

    name = raw.lower().strip()
    name = re.sub(r"[^a-z0-9]+", "-", name)
    return name.strip("-")


def _infer_difficulty(criteria_count: int) -> str:
    if criteria_count <= _EASY_THRESHOLD:
        return "easy"
    if criteria_count >= _HARD_THRESHOLD:
        return "hard"
    return "medium"


def _task_timeouts(criteria_count: int) -> tuple[int, int]:
    agent_timeout = max(1800, criteria_count * 30)
    verifier_timeout = max(300, criteria_count * 15)
    return agent_timeout, verifier_timeout


def _criteria_with_files_aliases(criteria: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for criterion in criteria:
        item = dict(criterion)
        if "files" not in item and isinstance(item.get("deliverables"), list):
            item["files"] = list(item["deliverables"])
        normalized.append(item)
    return normalized


def _build_task_toml(
    task_name: str,
    work_type: str,
    tags: list[str],
    criteria_count: int,
) -> str:
    """Generate task.toml content for a single task."""
    difficulty = _infer_difficulty(criteria_count)
    agent_timeout, verifier_timeout = _task_timeouts(criteria_count)

    tag_list = ", ".join(f'"{t}"' for t in tags)

    return textwrap.dedent(f"""\
        version = "1.0"

        [task]
        name = "{task_name}"

        [metadata]
        author_name = "Harvey AI"
        author_email = "labs@harvey.ai"
        difficulty = "{difficulty}"
        category = "legal-{work_type}"
        tags = [{tag_list}]

        [agent]
        timeout_sec = {agent_timeout}

        [verifier]
        timeout_sec = {verifier_timeout}

        [verifier.env]
        ANTHROPIC_API_KEY = "${{ANTHROPIC_API_KEY}}"

        [environment]
        build_timeout_sec = 600
        cpus = 1
        memory_mb = 4096
        storage_mb = 20480
    """)


def _build_task_md(
    task_name: str,
    title: str,
    instructions: str,
    deliverables: dict[str, str] | list[str],
    work_type: str,
    tags: list[str],
    criteria_count: int,
    task_id: str,
) -> str:
    """Generate native task.md content for a single task."""
    difficulty = _infer_difficulty(criteria_count)
    agent_timeout, verifier_timeout = _task_timeouts(criteria_count)
    instruction = _build_instruction_md(
        title,
        instructions,
        deliverables,
        work_type,
    ).strip()
    frontmatter: dict[str, Any] = {
        "schema_version": "1.3",
        "task": {
            "name": task_name,
        },
        "metadata": {
            "author_name": "Harvey AI",
            "author_email": "labs@harvey.ai",
            "difficulty": difficulty,
            "category": f"legal-{work_type}",
            "tags": tags,
        },
        "agent": {
            "timeout_sec": agent_timeout,
        },
        "verifier": {
            "timeout_sec": verifier_timeout,
            "env": {
                "ANTHROPIC_API_KEY": "${ANTHROPIC_API_KEY}",
            },
        },
        "environment": {
            "build_timeout_sec": 600,
            "cpus": 1,
            "memory_mb": 4096,
            "storage_mb": 20480,
        },
        "benchflow": {
            "document_version": "0.3",
            "source": {
                "benchmark": "Harvey LAB",
                "task_id": task_id,
                "work_type": work_type,
                "criteria_count": criteria_count,
            },
            "oracle": {
                "evidence": "oracle/README.md",
                "static_solution": False,
            },
            "verifier": {
                "spec": "verifier/verifier.md",
                "rubric": "verifier/rubrics/rubric.json",
                "entrypoint": "verifier/test.sh",
                "judge_model": "claude-sonnet-4-6",
                "implementation": {
                    "type": "script",
                    "judge": "llm-as-judge",
                    "outputs": {
                        "reward_json": "/logs/verifier/reward.json",
                        "reward_details": "/logs/verifier/reward-details.json",
                    },
                },
            },
        },
    }
    return render_task_md(frontmatter, instruction)


def _build_instruction_md(
    title: str,
    instructions: str,
    deliverables: dict[str, str] | list[str],
    work_type: str,
) -> str:
    """Generate instruction.md for the agent."""
    lines = [f"# {title}", ""]

    lines.append(instructions)
    lines.append("")

    if deliverables:
        lines.append("## Expected Deliverables")
        lines.append("")
        if isinstance(deliverables, dict):
            filenames = list(deliverables.values())
        elif isinstance(deliverables, list):
            filenames = list(deliverables)
        else:
            filenames = []
        for filename in filenames:
            lines.append(f"- `{filename}`")
        lines.append("")

    lines.append("## Workspace Layout")
    lines.append("")
    lines.append("- Input documents are in `documents/` (read-only).")
    lines.append("- Write deliverables to `output/`.")
    lines.append("- Use the `read` tool for .docx, .xlsx, .pptx, .pdf files.")
    lines.append("")

    return "\n".join(lines)


def _build_dockerfile(task_id: str, *, include_rubric: bool = True) -> str:
    """Generate a Dockerfile that sets up the evaluation environment.

    Uses a digest-pinned Python base image for reproducibility.
    """
    rubric_copy = ""
    if include_rubric:
        rubric_copy = (
            "\n# Copy rubric for the verifier\nCOPY rubric.json /app/rubric.json\n"
        )

    return textwrap.dedent(f"""\
        # Pinned by digest for reproducibility.
        FROM python:3.13-slim@sha256:dc1546eefcbe8caaa1f004f16ab76b204b5e1dbd58ff81b899f21cd40541232f

        RUN apt-get update -qq && apt-get install -y -qq \\
            pandoc \\
            curl \\
            && rm -rf /var/lib/apt/lists/*

        RUN pip install --no-cache-dir \\
            pdfplumber \\
            openpyxl \\
            python-docx \\
            python-pptx \\
            markitdown \\
            pandas \\
            anthropic

        WORKDIR /app

        # Copy task documents
        COPY documents/ /app/documents/
        {rubric_copy}

        # Create output directory (matches Harvey LAB's /workspace/output)
        RUN mkdir -p /app/output /logs/verifier /logs/agent /logs/artifacts
    """)


def _build_test_sh() -> str:
    """Generate test.sh — the verifier entry point."""
    return textwrap.dedent("""\
        #!/bin/bash
        set -euo pipefail

        verifier_log="${BENCHFLOW_VERIFIER_LOG:-/logs/verifier/verifier.log}"
        mkdir -p "$(dirname "$verifier_log")"
        exec > >(tee "$verifier_log") 2>&1

        VERIFIER_DIR="${BENCHFLOW_VERIFIER_DIR:-/verifier}"
        LEGACY_TESTS_DIR="${BENCHFLOW_LEGACY_TESTS_DIR:-/tests}"
        if [ ! -f "$VERIFIER_DIR/evaluate.py" ] && [ -f "$LEGACY_TESTS_DIR/evaluate.py" ]; then
            VERIFIER_DIR="$LEGACY_TESTS_DIR"
        fi

        if [ -n "${BENCHFLOW_RUBRIC_JSON:-}" ]; then
            rubric_file="$BENCHFLOW_RUBRIC_JSON"
        elif [ -f "$VERIFIER_DIR/rubrics/rubric.json" ]; then
            rubric_file="$VERIFIER_DIR/rubrics/rubric.json"
        else
            rubric_file=/app/rubric.json
        fi
        reward_file="${BENCHFLOW_REWARD_TEXT:-/logs/verifier/reward.txt}"
        reward_json="${BENCHFLOW_REWARD_JSON:-/logs/verifier/reward.json}"
        details_json="${BENCHFLOW_REWARD_DETAILS_JSON:-/logs/verifier/reward-details.json}"
        mkdir -p "$(dirname "$reward_file")" "$(dirname "$reward_json")" "$(dirname "$details_json")"

        if [ -n "${BENCHFLOW_OUTPUT_DIR:-}" ]; then
            OUTPUT_DIR="$BENCHFLOW_OUTPUT_DIR"
        elif [ -d /app/output ] && [ "$(ls -A /app/output 2>/dev/null)" ]; then
            OUTPUT_DIR=/app/output
        else
            OUTPUT_DIR=/app
        fi

        python3 "$VERIFIER_DIR/evaluate.py" \\
            --rubric "$rubric_file" \\
            --output-dir "$OUTPUT_DIR" \\
            --reward-file "$reward_file"

        python3 - "$reward_file" "$reward_json" "$details_json" <<'PY'
        from __future__ import annotations

        import json
        import sys
        from pathlib import Path

        reward_path = Path(sys.argv[1])
        reward_json_path = Path(sys.argv[2])
        details_json_path = Path(sys.argv[3])
        reward = float(reward_path.read_text().strip())
        reward_json_path.write_text(json.dumps({"reward": reward}, indent=2) + "\\n")

        evaluation_details = reward_path.parent / "evaluation_details.json"
        if evaluation_details.exists():
            details = json.loads(evaluation_details.read_text())
        else:
            details = {"source": "harvey-lab-llm-judge"}
        details.setdefault("reward", reward)
        details.setdefault("score", reward)
        details_json_path.write_text(json.dumps(details, indent=2) + "\\n")
        PY
    """)


def _build_evaluate_py() -> str:
    """Generate evaluate.py — LLM-as-judge scoring using Claude (Anthropic)."""
    return textwrap.dedent('''\
        """LLM-as-judge verifier for Harvey LAB tasks.

        Reads the rubric (criteria) from rubric.json, collects agent
        deliverables from the output directory, and grades each criterion
        using an LLM judge. Writes the aggregate reward (0.0-1.0) to the
        reward file.
        """

        import argparse
        import json
        import os
        import re
        import string
        import subprocess
        import sys
        import time
        from pathlib import Path

        def read_file_as_text(path: Path) -> str:
            """Read a file and return its content as plain text."""
            suffix = path.suffix.lower()
            try:
                if suffix == ".docx":
                    result = subprocess.run(
                        ["pandoc", str(path), "-t", "markdown",
                         "--wrap=none", "--track-changes=accept"],
                        capture_output=True, text=True, timeout=30,
                    )
                    if result.returncode != 0:
                        return f"(pandoc error: {result.stderr[:200]})"
                    return result.stdout
                if suffix == ".xlsx":
                    from openpyxl import load_workbook

                    wb = load_workbook(str(path), data_only=True)
                    parts = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        parts.append(f"=== Sheet: {sheet_name} ===")
                        for row in ws.iter_rows(values_only=True):
                            parts.append("\\t".join(
                                str(c) if c is not None else "" for c in row
                            ))
                    return "\\n".join(parts)
                if suffix == ".pptx":
                    from markitdown import MarkItDown
                    md = MarkItDown()
                    result = md.convert(str(path))
                    return result.text_content
                if suffix == ".pdf":
                    import pdfplumber

                    parts = []
                    with pdfplumber.open(path) as pdf:
                        for page in pdf.pages:
                            text = page.extract_text()
                            if text:
                                parts.append(text)
                    return "\\n".join(parts)
                return path.read_text(encoding="utf-8")
            except Exception as e:
                return f"(error reading {path.name}: {e})"


        VERDICT_PROMPT = string.Template("""You are evaluating a legal AI agent\'s work product against a specific quality criterion.

        ## Task
        $task_description

        ## Agent\'s Output
        $agent_output

        ## Criterion
        **$criterion_title**

        $match_criteria

        ## Instructions
        Evaluate the agent\'s output against the criterion above.
        - **PASS**: The agent\'s output satisfies the criterion as described
        - **FAIL**: The agent\'s output does not satisfy the criterion as described

        Respond with JSON only:

        ```json
        {
          "verdict": "pass" or "fail",
          "reasoning": "Brief explanation"
        }
        ```
        """)


        def call_judge(prompt: str, retries: int = 3) -> str:
            """Call Anthropic Claude API and return the text response."""
            import anthropic

            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                raise RuntimeError("ANTHROPIC_API_KEY not set")

            client = anthropic.Anthropic(api_key=api_key)

            for attempt in range(retries):
                try:
                    response = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=2048,
                        messages=[{"role": "user", "content": prompt}],
                    )
                    return response.content[0].text
                except Exception as e:
                    if attempt < retries - 1:
                        time.sleep(2 ** attempt)
                        continue
                    raise RuntimeError(f"Anthropic API failed after {retries} attempts: {e}")


        def parse_verdict(text: str) -> dict:
            """Extract JSON verdict from LLM response."""
            # Try code fences first
            match = re.search(r"```(?:json)?\\s*\\n?(.*?)\\n?```", text, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group(1).strip())
                except json.JSONDecodeError:
                    pass
            # Try balanced braces
            for i, ch in enumerate(text):
                if ch == "{":
                    depth = 0
                    for j in range(i, len(text)):
                        if text[j] == "{":
                            depth += 1
                        elif text[j] == "}":
                            depth -= 1
                        if depth == 0:
                            try:
                                return json.loads(text[i:j + 1])
                            except json.JSONDecodeError:
                                break
            raise ValueError(f"Could not parse verdict from: {text[:300]}")


        def judge_criterion(
            criterion: dict,
            task_title: str,
            deliverable_texts: dict[str, str],
        ) -> dict:
            """Grade a single criterion against deliverable content."""
            # Build the agent output context from relevant deliverables
            criterion_deliverables = criterion.get("deliverables", [])
            if criterion_deliverables:
                def _stem(name: str) -> str:
                    return Path(name).stem.lower()

                expected_stems = {_stem(d) for d in criterion_deliverables}
                relevant = {k: v for k, v in deliverable_texts.items()
                            if _stem(k) in expected_stems
                            or any(d.lower() in k.lower() for d in criterion_deliverables)}
            else:
                relevant = deliverable_texts

            if not relevant:
                return {
                    "id": criterion["id"],
                    "title": criterion["title"],
                    "verdict": "fail",
                    "reasoning": "No matching deliverable files found.",
                }

            agent_output = "\\n\\n".join(
                f"--- {name} ---\\n{content[:15000]}"
                for name, content in relevant.items()
            )

            try:
                prompt = VERDICT_PROMPT.safe_substitute(
                    task_description=task_title,
                    agent_output=agent_output,
                    criterion_title=criterion["title"],
                    match_criteria=criterion["match_criteria"],
                )
                response_text = call_judge(prompt)
                verdict = parse_verdict(response_text)
                return {
                    "id": criterion["id"],
                    "title": criterion["title"],
                    "verdict": verdict.get("verdict", "fail"),
                    "reasoning": verdict.get("reasoning", ""),
                }
            except Exception as e:
                return {
                    "id": criterion["id"],
                    "title": criterion["title"],
                    "verdict": "fail",
                    "reasoning": f"Judge error: {e}",
                }


        def find_deliverables(output_dir: Path) -> dict[str, str]:
            """Find and read all deliverable files in the output directory."""
            texts = {}
            for f in sorted(output_dir.iterdir()):
                if f.is_file() and f.name != "rubric.json" and not f.name.startswith("."):
                    if f.stat().st_size > 50_000_000:  # skip files > 50 MB
                        continue
                    if f.suffix.lower() in (".docx", ".xlsx", ".pptx", ".pdf",
                                             ".md", ".txt", ".json", ".csv"):
                        texts[f.name] = read_file_as_text(f)
            return texts


        def main():
            parser = argparse.ArgumentParser()
            parser.add_argument("--rubric", required=True, help="Path to rubric.json")
            parser.add_argument("--output-dir", required=True, help="Agent output directory")
            parser.add_argument("--reward-file", required=True, help="Path to write reward")
            args = parser.parse_args()

            rubric_path = Path(args.rubric)
            output_dir = Path(args.output_dir)
            reward_file = Path(args.reward_file)

            rubric = json.loads(rubric_path.read_text())
            task_title = rubric.get("title", "Legal task")
            criteria = rubric.get("criteria", [])

            if not criteria:
                reward_file.write_text("0")
                return

            deliverable_texts = find_deliverables(output_dir)

            if not deliverable_texts:
                print("No deliverable files found in output directory.")
                reward_file.write_text("0")
                return

            print(f"Found {len(deliverable_texts)} deliverable(s): "
                  f"{list(deliverable_texts.keys())}")
            print(f"Evaluating {len(criteria)} criteria...")

            results = []
            for i, criterion in enumerate(criteria):
                print(f"  [{i+1}/{len(criteria)}] {criterion['id']}: "
                      f"{criterion['title'][:60]}...")
                result = judge_criterion(criterion, task_title, deliverable_texts)
                results.append(result)
                status = "PASS" if result["verdict"] == "pass" else "FAIL"
                print(f"    -> {status}: {result['reasoning'][:80]}")

            n_passed = sum(1 for r in results if r["verdict"] == "pass")
            n_total = len(results)
            reward = n_passed / n_total if n_total > 0 else 0.0

            print(f"\\nScore: {n_passed}/{n_total} ({reward:.1%})")

            reward_file.parent.mkdir(parents=True, exist_ok=True)
            reward_file.write_text(str(reward))

            # Write detailed results alongside reward
            details_path = reward_file.parent / "evaluation_details.json"
            details_path.write_text(json.dumps({
                "reward": reward,
                "score": reward,
                "n_passed": n_passed,
                "n_total": n_total,
                "results": results,
            }, indent=2))


        if __name__ == "__main__":
            main()
    ''')


def _build_verifier_md(task_name: str, title: str, criteria_count: int) -> str:
    frontmatter: dict[str, Any] = {
        "document_version": "0.3",
        "verifier": {
            "name": f"{_sanitize_name(task_name)}-verifier",
            "default_strategy": "deterministic",
            "strategies": {
                "deterministic": {
                    "type": "script",
                    "command": "./test.sh",
                },
                "llm_judge": {
                    "type": "llm-judge",
                    "model": "claude-sonnet-4-6",
                    "rubric": "rubrics/rubric.json",
                    "input_dir": "/app/output",
                    "context_file": "rubrics/context.md",
                },
            },
            "rubric": {
                "combine": "mean",
                "dimensions": {
                    "criteria_pass_rate": {
                        "weight": 1.0,
                        "source": "deterministic",
                    },
                },
            },
            "outputs": {
                "reward_text": "/logs/verifier/reward.txt",
                "reward_json": "/logs/verifier/reward.json",
                "details_json": "/logs/verifier/reward-details.json",
                "aggregate_policy": {
                    "field": "reward",
                    "method": "mean",
                },
            },
        },
    }
    rendered_frontmatter = yaml.safe_dump(frontmatter, sort_keys=False)
    return (
        f"---\n{rendered_frontmatter}---\n\n## role:reviewer\n\n"
        f"Evaluate `{title}` with Harvey LAB's per-criterion legal rubric. "
        f"The task has {criteria_count} pass/fail criteria; reward is the "
        "fraction of criteria that the Claude judge marks as passing.\n"
    )


def _build_verifier_rubric_md(task_name: str, criteria_count: int) -> str:
    return f"""\
# Harvey LAB Verifier Rubric

Task: `{task_name}`

The machine-readable Harvey LAB criteria live in `rubrics/rubric.json`.

- Each criterion is judged independently as pass or fail.
- The verifier reads deliverables from `/app/output` and falls back to `/app`.
- Reward is `passed_criteria / {criteria_count}` when criteria are present.
"""


def _build_context_md(title: str, work_type: str) -> str:
    return f"""\
# Harvey LAB Judge Context

Title: {title}
Work type: {work_type}

Judge only the declared deliverables against the Harvey LAB rubric criteria.
Do not give credit for missing output files or unsupported assertions.
"""


def _build_oracle_readme(task_id: str, criteria_count: int) -> str:
    return f"""\
# Oracle Evidence

Harvey LAB task `{task_id}` does not ship static gold deliverables. The
benchmark's ground truth is its human-authored rubric: {criteria_count}
pass/fail criteria judged against the agent's legal work product.

The verifier package stores those criteria in `verifier/rubrics/rubric.json`
and computes reward as the fraction of criteria marked passing by the LLM judge.
"""


def _discover_tasks(harvey_root: Path) -> list[dict]:
    """Discover all Harvey LAB tasks from the tasks/ directory."""
    tasks_dir = harvey_root / "tasks"
    if not tasks_dir.exists():
        raise FileNotFoundError(f"Harvey LAB tasks dir not found: {tasks_dir}")

    discovered = []
    for task_json in sorted(tasks_dir.rglob("task.json")):
        rel = task_json.parent.relative_to(tasks_dir)
        task_id = str(rel).replace("\\", "/")

        with open(task_json) as f:
            config = json.load(f)

        docs_dir = task_json.parent / "documents"
        if not docs_dir.exists():
            print(f"  SKIP {task_id}: no documents/ directory", file=sys.stderr)
            continue

        discovered.append(
            {
                "task_id": task_id,
                "task_json_path": task_json,
                "docs_dir": docs_dir,
                "config": config,
            }
        )

    return discovered


def generate_task(
    task_info: dict,
    output_dir: Path,
    overwrite: bool = False,
    task_format: TaskFormat = "task-md",
) -> Path | None:
    """Generate a single BenchFlow task directory from a Harvey LAB task."""
    task_format = validate_task_output_format(task_format)
    task_id = task_info["task_id"]
    config = task_info["config"]
    docs_dir = task_info["docs_dir"]

    # Sanitize the task name for BenchFlow registry
    task_name = f"harvey-lab/{_sanitize_name(task_id)}"
    task_dir = output_dir / _sanitize_name(task_id)

    if task_dir.exists():
        if overwrite:
            shutil.rmtree(task_dir)
        else:
            ensure_existing_task_output_format(task_dir, task_format)
            return task_dir

    task_dir.mkdir(parents=True)

    # Extract fields from Harvey LAB config
    title = config.get("title", task_id)
    work_type = config.get("work_type", "analyze")
    tags = config.get("tags", [])
    instructions = config.get("instructions", "")
    deliverables = config.get("deliverables", {})
    criteria = config.get("criteria", [])

    # If instructions not inline, check for instructions.md
    if not instructions:
        instr_path = task_info["task_json_path"].parent / "instructions.md"
        if instr_path.exists():
            instructions = instr_path.read_text(encoding="utf-8")

    if task_format == "task-md":
        (task_dir / "task.md").write_text(
            _build_task_md(
                task_name,
                title,
                instructions,
                deliverables,
                work_type,
                tags,
                len(criteria),
                task_id,
            )
        )
        oracle_dir = task_dir / "oracle"
        oracle_dir.mkdir()
        (oracle_dir / "README.md").write_text(
            _build_oracle_readme(task_id, len(criteria))
        )
    else:
        (task_dir / "task.toml").write_text(
            _build_task_toml(task_name, work_type, tags, len(criteria))
        )
        (task_dir / "instruction.md").write_text(
            _build_instruction_md(title, instructions, deliverables, work_type)
        )

    env_dir = task_dir / "environment"
    env_dir.mkdir()
    (env_dir / "Dockerfile").write_text(
        _build_dockerfile(task_id, include_rubric=task_format == "legacy")
    )

    # Copy documents into environment for Docker COPY
    env_docs_dir = env_dir / "documents"
    shutil.copytree(docs_dir, env_docs_dir)

    rubric = {"title": title, "criteria": _criteria_with_files_aliases(criteria)}
    if task_format == "legacy":
        (env_dir / "rubric.json").write_text(json.dumps(rubric, indent=2))

    tests_dir = task_dir / verifier_dir_name(task_format)
    tests_dir.mkdir()
    test_sh = tests_dir / "test.sh"
    test_sh.write_text(_build_test_sh())
    test_sh.chmod(0o755)

    evaluate_py = tests_dir / "evaluate.py"
    evaluate_py.write_text(_build_evaluate_py())

    if task_format == "task-md":
        (tests_dir / "verifier.md").write_text(
            _build_verifier_md(task_name, title, len(criteria))
        )
        rubrics_dir = tests_dir / "rubrics"
        rubrics_dir.mkdir()
        (rubrics_dir / "rubric.json").write_text(json.dumps(rubric, indent=2))
        (rubrics_dir / "verifier.md").write_text(
            _build_verifier_rubric_md(task_name, len(criteria))
        )
        (rubrics_dir / "context.md").write_text(_build_context_md(title, work_type))

    return task_dir


def main():
    parser = argparse.ArgumentParser(
        description="Generate BenchFlow tasks from Harvey LAB benchmark.",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Directory to write generated BenchFlow task directories.",
    )
    parser.add_argument(
        "--harvey-root",
        default=str(_DEFAULT_HARVEY_ROOT),
        help="Path to the Harvey LAB repository root.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit the number of tasks to generate.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing task directories.",
    )
    parser.add_argument(
        "--task-ids",
        default=None,
        help="Comma-separated list of task IDs to generate (e.g., "
        "'corporate-ma/analyze-cim-deal-teaser/scenario-01').",
    )
    parser.add_argument(
        "--split",
        default="main",
        help=(
            "Which slice to generate. 'main' (default) = all 1,251 tasks; "
            "'parity' = first 50 tasks alphabetically (for parity experiments); "
            "'xlsx' = first 25 tasks with any .xlsx deliverable; "
            "otherwise interpreted as a practice-area filter (e.g. 'corporate-ma')."
        ),
    )
    parser.add_argument(
        "--task-format",
        choices=TASK_FORMATS,
        default="task-md",
        help="Output layout: legacy task.toml/instruction.md or native task.md.",
    )
    args = parser.parse_args()
    task_format = validate_task_output_format(args.task_format)

    harvey_root = Path(args.harvey_root)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Discovering Harvey LAB tasks in {harvey_root}/tasks/ ...")
    all_tasks = _discover_tasks(harvey_root)
    print(f"Found {len(all_tasks)} tasks.")

    # Apply --split filter
    split = args.split.lower()
    if split == "parity":
        all_tasks = all_tasks[:50]
        print(f"Split 'parity': first 50 tasks alphabetically → {len(all_tasks)}.")
    elif split == "xlsx":
        xlsx_tasks = []
        for t in all_tasks:
            deliverables = t["config"].get("deliverables", {})
            if isinstance(deliverables, dict):
                filenames = list(deliverables.values())
            elif isinstance(deliverables, list):
                filenames = list(deliverables)
            else:
                filenames = []
            if any(fn.endswith(".xlsx") for fn in filenames):
                xlsx_tasks.append(t)
        all_tasks = xlsx_tasks[:25]
        print(
            f"Split 'xlsx': first 25 tasks with .xlsx deliverables → {len(all_tasks)}."
        )
    elif split != "main":
        # Treat as a practice-area filter
        all_tasks = [t for t in all_tasks if t["task_id"].startswith(split)]
        print(f"Split '{split}' (practice area): {len(all_tasks)} tasks.")

    # Filter by task IDs if specified
    if args.task_ids:
        requested = {t.strip() for t in args.task_ids.split(",")}
        all_tasks = [t for t in all_tasks if t["task_id"] in requested]
        print(f"Filtered to {len(all_tasks)} requested tasks.")

    # Apply limit
    if args.limit:
        all_tasks = all_tasks[: args.limit]
        print(f"Limited to {len(all_tasks)} tasks.")

    # Generate
    generated = 0
    errors = 0
    for task_info in all_tasks:
        try:
            result = generate_task(
                task_info,
                output_dir,
                overwrite=args.overwrite,
                task_format=task_format,
            )
            if result:
                generated += 1
        except Exception as e:
            print(f"  ERROR {task_info['task_id']}: {e}", file=sys.stderr)
            errors += 1

    print(f"\nGenerated {generated} tasks ({errors} errors) in {output_dir}")


if __name__ == "__main__":
    main()
