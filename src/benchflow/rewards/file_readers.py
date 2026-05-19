"""Document text extraction utilities for judge evaluation.

Extracts plain text from common document formats. Dependencies are
optional — missing packages produce a short error string instead of
crashing the verifier.
"""

from __future__ import annotations

import logging
import subprocess
from pathlib import Path

logger = logging.getLogger(__name__)

# File extensions we know how to handle.
SUPPORTED_EXTENSIONS = frozenset(
    {".docx", ".xlsx", ".pptx", ".pdf", ".md", ".txt", ".json", ".csv"}
)


def read_file_as_text(path: Path) -> str:
    """Read *path* and return its content as plain text.

    For rich document formats the appropriate library is used; if that
    library is not installed a short ``(unsupported ...)`` placeholder is
    returned so the caller can still proceed.
    """
    suffix = path.suffix.lower()
    try:
        if suffix == ".docx":
            return _read_docx(path)
        if suffix == ".xlsx":
            return _read_xlsx(path)
        if suffix == ".pptx":
            return _read_pptx(path)
        if suffix == ".pdf":
            return _read_pdf(path)
        # Plain-text fallback (.md, .txt, .json, .csv, ...)
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        logger.warning("Error reading %s: %s", path.name, exc)
        return f"(error reading {path.name}: {exc})"


def find_deliverables(directory: Path) -> dict[str, str]:
    """Discover and read all deliverable files in *directory*.

    Skips files larger than 50 MB and hidden/internal files.
    """
    texts: dict[str, str] = {}
    if not directory.is_dir():
        return texts
    for f in sorted(directory.iterdir()):
        if not f.is_file():
            continue
        if f.name.startswith(".") or f.name == "rubric.json":
            continue
        if f.stat().st_size > 50_000_000:
            continue
        if f.suffix.lower() in SUPPORTED_EXTENSIONS:
            texts[f.name] = read_file_as_text(f)
    return texts


# ------------------------------------------------------------------
# Format-specific readers
# ------------------------------------------------------------------


def _read_docx(path: Path) -> str:
    """Read .docx via pandoc (preferred) or python-docx."""
    try:
        result = subprocess.run(
            [
                "pandoc",
                str(path),
                "-t",
                "markdown",
                "--wrap=none",
                "--track-changes=accept",
            ],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode == 0:
            return result.stdout
    except FileNotFoundError:
        pass  # pandoc not installed

    try:
        from docx import Document  # ty: ignore[unresolved-import]

        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)
    except ImportError:
        return f"(unsupported: {path.name} — install pandoc or python-docx)"


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook  # ty: ignore[unresolved-import]

        wb = load_workbook(str(path), data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                parts.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(parts)
    except ImportError:
        return f"(unsupported: {path.name} — install openpyxl)"


def _read_pptx(path: Path) -> str:
    try:
        from markitdown import MarkItDown  # ty: ignore[unresolved-import]

        md = MarkItDown()
        result = md.convert(str(path))
        return result.text_content
    except ImportError:
        return f"(unsupported: {path.name} — install markitdown)"


def _read_pdf(path: Path) -> str:
    try:
        import pdfplumber  # ty: ignore[unresolved-import]

        parts: list[str] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    parts.append(text)
        return "\n".join(parts)
    except ImportError:
        return f"(unsupported: {path.name} — install pdfplumber)"
